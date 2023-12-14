import sys
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QTextEdit, QPushButton, QVBoxLayout, QFileDialog, QMessageBox, QProgressBar, QTimeEdit
from PyQt5.QtCore import QTimer, pyqtSignal
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
import smtplib
import openpyxl
import threading
import time

class EmailSenderApp(QWidget):
    send_finished = pyqtSignal(int)

    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle('邮件发送程序')
        self.setGeometry(100, 100, 400, 400)

        self.label_sender = QLabel('发件人邮箱:')
        self.sender_email = QLineEdit()

        self.label_password = QLabel('授权码:')
        self.password = QLineEdit()
        self.password.setEchoMode(QLineEdit.Password)

        self.label_subject = QLabel('邮件主题:')
        self.subject = QLineEdit()

        self.label_body = QLabel('邮件正文:')
        self.body = QTextEdit()

        self.label_file = QLabel('附件:')
        self.file_path = QLineEdit()
        self.browse_file_button = QPushButton('浏览')
        self.browse_file_button.clicked.connect(self.browse_file)

        self.label_recipients = QLabel('收件人表格:')
        self.recipients_file_path = QLineEdit()
        self.browse_recipients_button = QPushButton('浏览')
        self.browse_recipients_button.clicked.connect(self.browse_recipients_file)

        self.label_daily_limit = QLabel('每天发送数量:')
        self.daily_limit = QLineEdit()

        self.label_send_time = QLabel('选择发送时间:')
        self.send_time = QTimeEdit()
        self.send_time.setDisplayFormat("HH:mm")

        self.send_button = QPushButton('开始发送')
        self.send_button.clicked.connect(self.start_sending)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)

        layout = QVBoxLayout()
        layout.addWidget(self.label_sender)
        layout.addWidget(self.sender_email)
        layout.addWidget(self.label_password)
        layout.addWidget(self.password)
        layout.addWidget(self.label_subject)
        layout.addWidget(self.subject)
        layout.addWidget(self.label_body)
        layout.addWidget(self.body)
        layout.addWidget(self.label_file)
        layout.addWidget(self.file_path)
        layout.addWidget(self.browse_file_button)
        layout.addWidget(self.label_recipients)
        layout.addWidget(self.recipients_file_path)
        layout.addWidget(self.browse_recipients_button)
        layout.addWidget(self.label_daily_limit)
        layout.addWidget(self.daily_limit)
        layout.addWidget(self.label_send_time)
        layout.addWidget(self.send_time)
        layout.addWidget(self.send_button)
        layout.addWidget(self.progress_bar)

        self.setLayout(layout)

        self.send_finished.connect(self.show_finished_message)

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, '选择附件', '', 'All Files (*)')
        self.file_path.setText(file_path)

    def browse_recipients_file(self):
        recipients_file_path, _ = QFileDialog.getOpenFileName(self, '选择收件人表格', '', 'Excel Files (*.xlsx);;All Files (*)')
        self.recipients_file_path.setText(recipients_file_path)

    def send_email(self, sender, password, recipient, subject, body, attachment=None):
        try:
            msg = MIMEMultipart()
            msg['From'] = sender
            msg['To'] = recipient
            msg['Subject'] = subject

            msg.attach(MIMEText(body, 'plain'))

            if attachment:
                with open(attachment, 'rb') as file:
                    attachment_part = MIMEApplication(file.read(), Name=attachment)
                    attachment_part['Content-Disposition'] = f'attachment; filename={attachment}'
                    msg.attach(attachment_part)

            # Automatically detect email provider and adjust settings
            if sender.endswith('@qq.com'):
                smtp_server = 'smtp.qq.com'
                smtp_port = 465
                use_ssl = True
            elif sender.endswith('@163.com'):
                smtp_server = 'smtp.163.com'
                smtp_port = 465
                use_ssl = True
            elif sender.endswith('@gmail.com'):
                smtp_server = 'smtp.gmail.com'
                smtp_port = 587
                use_ssl = False
            else:
                # Add more conditions for other email providers if needed
                raise ValueError('Unsupported email provider')

            with smtplib.SMTP_SSL(smtp_server, smtp_port) if use_ssl else smtplib.SMTP(smtp_server, smtp_port) as server:
                if not use_ssl:
                    server.starttls()  # Use STARTTLS only if the server supports it
                server.login(sender, password)
                server.sendmail(sender, recipient, msg.as_string())
            return True
        except Exception as e:
            print(f"Error sending email to {recipient}: {e}")
            return False

    def start_sending(self):
        sender = self.sender_email.text()
        password = self.password.text()
        subject = self.subject.text()
        body = self.body.toPlainText()
        attachment = self.file_path.text()
        recipients_file_path = self.recipients_file_path.text()
        daily_limit = int(self.daily_limit.text())
        send_time = self.send_time.time()

        if not sender or not password or not subject or not body or not recipients_file_path or not daily_limit:
            QMessageBox.critical(self, '错误', '请填写所有字段和选择收件人表格')
            return

        try:
            recipients_wb = openpyxl.load_workbook(recipients_file_path)
            sheet = recipients_wb.active
            recipients = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row + 1)]
        except Exception as e:
            QMessageBox.critical(self, '错误', f'无法读取收件人表格: {e}')
            return

        total_recipients = len(recipients)
        emails_sent = 0

        def send_emails():
            nonlocal emails_sent
            for recipient in recipients:
                current_time = time.localtime()
                if current_time.tm_hour == send_time.hour() and current_time.tm_min == send_time.minute():
                    if self.send_email(sender, password, recipient, subject, body, attachment):
                        emails_sent += 1
                        time.sleep(1)  # Add a delay to avoid being flagged as spam
                        if emails_sent >= daily_limit:
                            break

            self.send_finished.emit(emails_sent)

        thread = threading.Thread(target=send_emails)
        thread.start()

        # Show progress bar while sending emails
        self.progress_bar.setRange(0, daily_limit)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)

        # Check progress every second
        timer = QTimer(self)
        timer.timeout.connect(lambda: self.progress_bar.setValue(emails_sent))
        timer.start(1000)

    def show_finished_message(self, emails_sent):
        self.progress_bar.setVisible(False)
        QMessageBox.information(self, '发送完成', f'成功发送 {emails_sent} 封邮件')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = EmailSenderApp()
    window.show()
    sys.exit(app.exec_())
